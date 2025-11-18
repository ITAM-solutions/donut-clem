"""
File name: excel_analysis
Type: Standalone Script
Author: Fran Moreno
Last Updated: 5/29/2025
Version: 1.0
Description:

This script is designed to prepare the JDE dataset called C&O, containing archive invoices and contracts, together with
some of their data extractions in Excel format. The intention is to find all the files that can be used to train
Donut. If those files do have an Excel data extraction, it will be used to take a first step in the tagging stage.
"""
import datetime
import shutil
import tqdm
import openpyxl
import glob
import os
from pathlib import Path
import math
import json

DATASET_PATH = Path(r'C:\Users\FranMoreno\datasets\c_and_o')
OUTPUT_PATH = Path(r'C:\Users\FranMoreno\datasets\c_and_o_outputs_v2')


def convert_size(size_bytes):
    if size_bytes == 0:
        return "0B"
    size_name = ("B", "KB", "MB", "GB", "TB", "PB", "EB", "ZB", "YB")
    i = int(math.floor(math.log(size_bytes, 1024)))
    p = math.pow(1024, i)
    s = round(size_bytes / p, 2)
    return "%s %s" % (s, size_name[i])


def get_all_excel_files() -> list[str]:
    return [y for x in os.walk(DATASET_PATH) for y in glob.glob(os.path.join(x[0], '*.xlsx'))]

def get_all_pdf_files(just_names) -> list[str]:
    if just_names:
        return [Path(y).name for x in os.walk(DATASET_PATH) for y in glob.glob(os.path.join(x[0], '*.pdf'))]
    return [y for x in os.walk(DATASET_PATH) for y in glob.glob(os.path.join(x[0], '*.pdf'))]


def get_sheet_info():
    excel_files = get_all_excel_files()
    entitlement_count = 0
    for file in excel_files:
        wb = openpyxl.load_workbook(filename=file, read_only=True, data_only=True)
        sheets = wb.sheetnames
        for sheet in sheets:
            if 'entitlement' in sheet.lower():
                entitlement_count += 1
                file_path = Path(file)
                print(f'FILE: {file} [{sheet}]')

        wb.close()
    print(f'Total excel files: {len(excel_files)}')
    print(f'Files with Entitlement sheet: {entitlement_count}')


def find_source_documents_from_excel_sheets():
    excel_files = get_all_excel_files()
    pdf_files = get_all_pdf_files(True)
    target_header = 'Source document'
    for file in excel_files:
        wb = openpyxl.load_workbook(filename=file, read_only=True, data_only=True)
        sheets = wb.sheetnames
        for sheet in sheets:
            ws = wb[sheet]
            rows = ws.iter_rows(values_only=True)

            try:
                header = next(rows)
                if target_header in header:
                    col_idx = header.index(target_header)
                    values = [row[col_idx] for row in rows if row and len(row) > col_idx]

                    # Now check if the values exist in the pdf list
                    for value in values:
                        if value in pdf_files:
                            print('True!')

            except Exception:
                continue


def convert_if_date(value):
    if isinstance(value, datetime.datetime):
        return value.strftime("%d/%m/%Y")
    return value


def find_entitlement_files():
    # Finds all Entitlement PDF files, including archive directories.
    entitlement_files = [y for x in os.walk(DATASET_PATH) for y in glob.glob(os.path.join(x[0], 'Entitlement*/**/*.pdf'), recursive=True)]

    def gather_by_provider(file_list: list[str]) -> dict[Path, list[Path]]:
        provider_mapping = dict()
        for file in file_list:
            file_path = Path(file)
            parts = file_path.parts
            provider_idx = parts.index('c_and_o') + 1
            provider_path = Path(*parts[:provider_idx + 1])
            file_path_relative = Path(*parts[provider_idx + 1:])
            if provider_path in provider_mapping:
                provider_mapping[provider_path].append(file_path_relative)
            else:
                provider_mapping[provider_path] = [file_path_relative]
        return provider_mapping

    entitlements_per_provider = gather_by_provider(entitlement_files)

    def search_in_corresponding_excels(files_per_provider: dict[Path, list[Path]]):
        excel_to_pdf_mapping = dict()

        for provider_path, files in tqdm.tqdm(files_per_provider.items()):
            excel_files = [y for x in os.walk(provider_path) for y in glob.glob(os.path.join(x[0], '*.xlsx'), recursive=True)]
            file_names = [file.stem for file in files]

            target_headers = ['Source document', 'Agreement name']

            # Search the files names in the excel to determine a pdf-excel relationship.
            for excel in excel_files:
                print(excel)
                wb = openpyxl.load_workbook(filename=excel, read_only=True, data_only=True)
                sheets = wb.sheetnames

                for sheet in sheets:
                    if 'entitlement' not in sheet.lower():
                        continue

                    ws = wb[sheet]
                    rows = ws.iter_rows(values_only=True, max_row=2000)

                    try:
                        headers_row = next(rows)

                        data_rows = [row for row in rows if any(cell is not None for cell in row)]
                        print('data loaded!')
                        for target_header in target_headers:
                            if target_header in headers_row:
                                col_idx = headers_row.index(target_header)
                                column_values = [row[col_idx] for row in data_rows]

                                # Check if any of the found values exist in the pdf list
                                for value in column_values:
                                    if value in file_names:
                                        file_idx = file_names.index(value)
                                        file_path_absolute = provider_path / files[file_idx]

                                        fields = list()
                                        for data_row in data_rows:
                                            if value in data_row:
                                                fields.append({f_name: convert_if_date(f_val) for f_name, f_val in zip(headers_row, data_row)})

                                        # Save data to json
                                        file_data = {
                                            "metadata": {
                                                "full_path": str(file_path_absolute),
                                                "excel_path": excel,
                                                "sheet_name": sheet,
                                                "mapped_in_column": target_header
                                            },
                                            "values": fields
                                        }

                                        output_path = OUTPUT_PATH / provider_path.name / (files[file_idx].stem + '.json')
                                        if not output_path.parent.exists():
                                            os.makedirs(output_path.parent)
                                        if output_path.exists():
                                            continue
                                        with open(output_path, 'w') as f:
                                            json.dump(file_data, f, indent=2)

                    except StopIteration:
                        # Target header not found. This excel does not contain linking info to any PDF.
                        continue

                wb.close()

        return excel_to_pdf_mapping

    search_in_corresponding_excels(entitlements_per_provider)


def copy_useful_files():
    # Finds all Entitlement PDF files, including archive directories.
    entitlement_files = [Path(y) for x in os.walk(DATASET_PATH) for y in
        glob.glob(os.path.join(x[0], 'Entitlement*/**/*.pdf'), recursive=True)]


    # Get the name of all the json files generated with 'find_entitlement_files()'
    files_with_useful_output = [Path(y) for x in os.walk(OUTPUT_PATH) for y in
        glob.glob(os.path.join(x[0], '**/*.json'))]

    counter = 0
    entitlement_names = [f.stem for f in entitlement_files]
    for output_file in files_with_useful_output:
        if output_file.stem not in entitlement_names:
            counter += 1


    files_with_useful_output_names = [i.stem for i in files_with_useful_output]

    for org_file in tqdm.tqdm(entitlement_files):
        if org_file.stem in files_with_useful_output_names:
            idx = files_with_useful_output_names.index(org_file.stem)
            shutil.copy(org_file, files_with_useful_output[idx].parent / org_file.name)
        else:
            parts = org_file.parts
            provider_idx = parts.index('c_and_o') + 1
            provider_path = Path(*parts[:provider_idx + 1])
            output_path = OUTPUT_PATH / provider_path.name / org_file.name

            if not output_path.parent.exists():
                os.makedirs(output_path.parent)

            shutil.copy(org_file, output_path)

def main():
    # find_source_documents_from_excel_sheets()

    # ----
    find_entitlement_files()
    copy_useful_files()
    # ----
    pass

if __name__ == '__main__':
    main()
